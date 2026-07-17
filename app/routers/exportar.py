from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.dependencies import verificar_token
from app.models.personal import Personal

router = APIRouter(
    prefix="/admin/exportar",
    tags=["Exportar"],
    dependencies=[Depends(verificar_token)],
)

# ── Estilos ────────────────────────────────────────────────
AZUL_HEADER = "1E3A8A"
AZUL_SECCION = "2563EB"
AMARILLO = "FEF08A"
GRIS_FILA = "F8FAFC"
BLANCO = "FFFFFF"


def estilo_header(cell, color=AZUL_HEADER):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borde_fino()


def estilo_subheader(cell):
    cell.font = Font(bold=True, color="1E3A8A", size=9)
    cell.fill = PatternFill("solid", fgColor=AMARILLO)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borde_fino()


def estilo_dato(cell, fila_par=False):
    cell.font = Font(size=9)
    cell.fill = PatternFill("solid", fgColor=GRIS_FILA if fila_par else BLANCO)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = borde_fino()


def borde_fino():
    lado = Side(style="thin", color="CBD5E1")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def v(val):
    """Convierte None a cadena vacía."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Sí" if val else "No"
    return str(val)


def aplicar_estilos_hoja(ws, headers, col_widths):
    """Aplica estilos a los encabezados y ajusta anchos."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        estilo_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


@router.get(
    "/excel",
    summary="Exportar todo el personal a Excel",
    response_class=StreamingResponse,
)
async def exportar_excel(
    db: AsyncSession = Depends(get_db),
):
    # ── Cargar todo el personal con relaciones ─────────────
    resultado = await db.execute(
        select(Personal)
        .options(
            selectinload(Personal.datos_laborales),
            selectinload(Personal.familiares),
            selectinload(Personal.formacion_academica),
            selectinload(Personal.otros_estudios),
            selectinload(Personal.experiencia_laboral),
            selectinload(Personal.experiencia_docente),
            selectinload(Personal.otras_instituciones),
            selectinload(Personal.reconocimientos),
        )
        .order_by(Personal.apellido_paterno, Personal.apellido_materno)
    )
    personal_list = resultado.scalars().unique().all()

    # ── Crear workbook ─────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # ══════════════════════════════════════════════════════
    # HOJA 1: RESUMEN DEL PERSONAL
    # ══════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Resumen Personal")

    headers_resumen = [
        "N°",
        "APELLIDOS Y NOMBRES",
        "DNI",
        "SEXO",
        "FECHA NACIMIENTO",
        "CELULAR",
        "EMAIL PERSONAL",
        "DEPENDENCIA",
        "CARGO",
        "CONDICIÓN",
        "TIPO PERSONAL",
        "RÉGIMEN",
        "FECHA INGRESO",
        "EMAIL INSTITUCIONAL",
        "BANCO",
        "N° CUENTA",
        "CCI",
        "SISTEMA PENSIÓN",
        "AFP",
        "FOTO",
    ]
    col_widths_resumen = [
        5,
        35,
        12,
        10,
        14,
        14,
        28,
        30,
        25,
        12,
        15,
        20,
        14,
        28,
        12,
        20,
        22,
        14,
        12,
        8,
    ]
    aplicar_estilos_hoja(ws1, headers_resumen, col_widths_resumen)

    for idx, p in enumerate(personal_list, 1):
        l = p.datos_laborales
        fila_par = idx % 2 == 0

        # Resolver régimen
        regimen = ""
        if l:
            if l.regimen_dl276:
                regimen = f"DL 276 — {l.regimen_dl276}"
            elif l.regimen_cas:
                regimen = f"CAS — {l.regimen_cas}"
            elif l.regimen_ordinario:
                regimen = f"Ord. — {l.regimen_ordinario}"
            elif l.regimen_contratado:
                regimen = f"Cont. — {l.regimen_contratado}"
            elif l.regimen_otros:
                regimen = l.regimen_otros

        fila = [
            idx,
            f"{v(p.apellido_paterno)} {v(p.apellido_materno)}, {v(p.nombres)}",
            v(p.dni),
            v(p.sexo),
            v(p.fecha_nacimiento),
            v(p.celular),
            v(p.email_personal_1),
            v(l.dependencia) if l else "",
            v(l.cargo) if l else "",
            v(l.condicion) if l else "",
            v(l.tipo_personal) if l else "",
            regimen,
            v(l.fecha_ingreso) if l else "",
            v(l.email_institucional) if l else "",
            v(p.banco),
            v(p.cuenta_numero),
            v(p.cuenta_cci),
            v(p.sistema_pension),
            v(p.afp_nombre),
            "Sí" if p.foto_url else "No",
        ]
        row_num = idx + 1
        for col_idx, valor in enumerate(fila, 1):
            cell = ws1.cell(row=row_num, column=col_idx, value=valor)
            estilo_dato(cell, fila_par)
        ws1.row_dimensions[row_num].height = 18

    # Total al final
    row_total = len(personal_list) + 2
    cell_total = ws1.cell(
        row=row_total, column=1, value=f"TOTAL: {len(personal_list)} registros"
    )
    cell_total.font = Font(bold=True, color="1E3A8A", size=9)
    cell_total.fill = PatternFill("solid", fgColor=AMARILLO)

    # ══════════════════════════════════════════════════════
    # HOJA 2: DETALLE COMPLETO
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle Completo")

    # Encabezado principal
    ws2.merge_cells("A1:Z1")
    title_cell = ws2["A1"]
    title_cell.value = (
        f"FICHA DE REGISTRO DE DATOS DEL PERSONAL — " f"UNAMBA {datetime.now().year}"
    )
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = PatternFill("solid", fgColor=AZUL_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    fila_actual = 2

    for idx, p in enumerate(personal_list, 1):
        l = p.datos_laborales
        fam = p.familiares or []
        fo = p.formacion_academica or []
        oe = p.otros_estudios or []
        el = p.experiencia_laboral or []
        ed = p.experiencia_docente or []
        oi = p.otras_instituciones
        re = p.reconocimientos or []

        # ── Separador por persona ──────────────────────
        ws2.merge_cells(f"A{fila_actual}:Z{fila_actual}")
        sep = ws2.cell(
            row=fila_actual,
            column=1,
            value=(
                f"  {idx}. "
                f"{v(p.apellido_paterno)} {v(p.apellido_materno)}, "
                f"{v(p.nombres)}   |   DNI: {v(p.dni)}"
            ),
        )
        sep.font = Font(bold=True, color="FFFFFF", size=10)
        sep.fill = PatternFill("solid", fgColor=AZUL_SECCION)
        sep.alignment = Alignment(vertical="center")
        ws2.row_dimensions[fila_actual].height = 22
        fila_actual += 1

        # ── DATOS PERSONALES ───────────────────────────
        campos_personales = [
            ("Apellido Paterno", v(p.apellido_paterno)),
            ("Apellido Materno", v(p.apellido_materno)),
            ("Nombres", v(p.nombres)),
            ("DNI", v(p.dni)),
            ("Sexo", v(p.sexo)),
            ("Fecha Nacimiento", v(p.fecha_nacimiento)),
            ("Estado Civil", v(p.estado_civil)),
            ("Departamento Nac.", v(p.nac_departamento)),
            ("Provincia Nac.", v(p.nac_provincia)),
            ("Distrito Nac.", v(p.nac_distrito)),
            ("Celular", v(p.celular)),
            ("Teléfono Fijo", v(p.telefono_fijo)),
            ("Email Personal 1", v(p.email_personal_1)),
            ("Email Personal 2", v(p.email_personal_2)),
            ("Dirección", v(p.dom_direccion)),
            ("Tipo Vivienda", v(p.tipo_vivienda)),
            ("RUC", v(p.ruc)),
            ("Lic. Conducir", v(p.licencia_conducir)),
            ("Afil. ESSALUD", v(p.afiliacion_essalud)),
            ("Grupo Sanguíneo", v(p.grupo_sanguineo)),
            ("Donador Órganos", v(p.donador_organos)),
            ("Banco", v(p.banco)),
            ("N° Cuenta", v(p.cuenta_numero)),
            ("CCI", v(p.cuenta_cci)),
            ("Denominación Prof.", v(p.denominacion_prof)),
            ("Colegio Prof.", v(p.colegio_prof_nombre)),
            ("N° Colegiatura", v(p.colegio_prof_numero)),
            ("Sistema Pensión", v(p.sistema_pension)),
            ("AFP", v(p.afp_nombre)),
            ("Discapacidad", v(p.tiene_discapacidad)),
            ("Serv. Militar", v(p.realizo_serv_militar)),
        ]

        # Sub-header Datos Personales
        ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
        sh = ws2.cell(row=fila_actual, column=1, value="DATOS PERSONALES")
        estilo_subheader(sh)
        fila_actual += 1

        # Datos en 4 columnas: label | valor | label | valor
        for i in range(0, len(campos_personales), 2):
            par = campos_personales[i : i + 2]
            c1 = ws2.cell(row=fila_actual, column=1, value=par[0][0])
            c2 = ws2.cell(row=fila_actual, column=2, value=par[0][1])
            c1.font = Font(bold=True, size=9)
            c1.fill = PatternFill("solid", fgColor="EFF6FF")
            c1.border = borde_fino()
            c2.font = Font(size=9)
            c2.border = borde_fino()
            if len(par) > 1:
                c3 = ws2.cell(row=fila_actual, column=3, value=par[1][0])
                c4 = ws2.cell(row=fila_actual, column=4, value=par[1][1])
                c3.font = Font(bold=True, size=9)
                c3.fill = PatternFill("solid", fgColor="EFF6FF")
                c3.border = borde_fino()
                c4.font = Font(size=9)
                c4.border = borde_fino()
            fila_actual += 1

        # ── DATOS LABORALES ────────────────────────────
        if l:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(row=fila_actual, column=1, value="DATOS LABORALES")
            estilo_subheader(sh)
            fila_actual += 1

            regimen = ""
            if l.regimen_dl276:
                regimen = f"DL 276 — {l.regimen_dl276}"
            elif l.regimen_cas:
                regimen = f"CAS — {l.regimen_cas}"
            elif l.regimen_ordinario:
                regimen = f"Ordinario — {l.regimen_ordinario}"
            elif l.regimen_contratado:
                regimen = f"Contratado — {l.regimen_contratado}"
            elif l.regimen_otros:
                regimen = l.regimen_otros

            campos_lab = [
                ("Dependencia", v(l.dependencia)),
                ("Cargo", v(l.cargo)),
                ("Condición", v(l.condicion)),
                ("Tipo Personal", v(l.tipo_personal)),
                ("Fecha Ingreso", v(l.fecha_ingreso)),
                ("Email Institucional", v(l.email_institucional)),
                ("Régimen", regimen),
                ("Dedicación", v(l.dedicacion)),
            ]
            for i in range(0, len(campos_lab), 2):
                par = campos_lab[i : i + 2]
                c1 = ws2.cell(row=fila_actual, column=1, value=par[0][0])
                c2 = ws2.cell(row=fila_actual, column=2, value=par[0][1])
                c1.font = Font(bold=True, size=9)
                c1.fill = PatternFill("solid", fgColor="EFF6FF")
                c1.border = borde_fino()
                c2.font = Font(size=9)
                c2.border = borde_fino()
                if len(par) > 1:
                    c3 = ws2.cell(row=fila_actual, column=3, value=par[1][0])
                    c4 = ws2.cell(row=fila_actual, column=4, value=par[1][1])
                    c3.font = Font(bold=True, size=9)
                    c3.fill = PatternFill("solid", fgColor="EFF6FF")
                    c3.border = borde_fino()
                    c4.font = Font(size=9)
                    c4.border = borde_fino()
                fila_actual += 1

        # ── FAMILIARES ─────────────────────────────────
        if fam:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(row=fila_actual, column=1, value=f"FAMILIARES ({len(fam)})")
            estilo_subheader(sh)
            fila_actual += 1

            headers_fam = ["Apellidos y Nombres", "Parentesco", "DNI", "F. Nacimiento"]
            for ci, h in enumerate(headers_fam, 1):
                c = ws2.cell(row=fila_actual, column=ci, value=h)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="64748B")
                c.border = borde_fino()
                c.alignment = Alignment(horizontal="center")
            fila_actual += 1

            for f in fam:
                datos_fam = [
                    f"{v(f.apellido_paterno)} {v(f.apellido_materno)}, {v(f.nombres)}",
                    v(f.parentesco),
                    v(f.dni),
                    v(f.fecha_nacimiento),
                ]
                for ci, val in enumerate(datos_fam, 1):
                    c = ws2.cell(row=fila_actual, column=ci, value=val)
                    c.font = Font(size=9)
                    c.border = borde_fino()
                fila_actual += 1

        # ── FORMACIÓN ACADÉMICA ────────────────────────
        if fo:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(
                row=fila_actual, column=1, value=f"FORMACIÓN ACADÉMICA ({len(fo)})"
            )
            estilo_subheader(sh)
            fila_actual += 1

            headers_fo = [
                "Nivel",
                "Centro de Estudios",
                "Grado Obtenido",
                "Fecha Conclusión",
            ]
            for ci, h in enumerate(headers_fo, 1):
                c = ws2.cell(row=fila_actual, column=ci, value=h)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="64748B")
                c.border = borde_fino()
                c.alignment = Alignment(horizontal="center")
            fila_actual += 1

            for f in fo:
                datos_fo = [
                    v(f.nivel),
                    v(f.centro_estudios),
                    v(f.grado_obtenido),
                    v(f.fecha_conclusion),
                ]
                for ci, val in enumerate(datos_fo, 1):
                    c = ws2.cell(row=fila_actual, column=ci, value=val)
                    c.font = Font(size=9)
                    c.border = borde_fino()
                fila_actual += 1

        # ── OTROS ESTUDIOS ─────────────────────────────
        if oe:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(
                row=fila_actual, column=1, value=f"OTROS ESTUDIOS ({len(oe)})"
            )
            estilo_subheader(sh)
            fila_actual += 1

            headers_oe = [
                "Tipo",
                "Nombre del Curso",
                "Centro de Estudios",
                "Duración (hrs)",
            ]
            for ci, h in enumerate(headers_oe, 1):
                c = ws2.cell(row=fila_actual, column=ci, value=h)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="64748B")
                c.border = borde_fino()
                c.alignment = Alignment(horizontal="center")
            fila_actual += 1

            for e in oe:
                datos_oe = [
                    v(e.tipo),
                    v(e.nombre_curso),
                    v(e.centro_estudios),
                    v(e.duracion_horas),
                ]
                for ci, val in enumerate(datos_oe, 1):
                    c = ws2.cell(row=fila_actual, column=ci, value=val)
                    c.font = Font(size=9)
                    c.border = borde_fino()
                fila_actual += 1

        # ── EXPERIENCIA LABORAL ────────────────────────
        if el:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(
                row=fila_actual, column=1, value=f"EXPERIENCIA LABORAL ({len(el)})"
            )
            estilo_subheader(sh)
            fila_actual += 1

            headers_el = ["Tipo", "Entidad", "Cargo", "Período"]
            for ci, h in enumerate(headers_el, 1):
                c = ws2.cell(row=fila_actual, column=ci, value=h)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="64748B")
                c.border = borde_fino()
                c.alignment = Alignment(horizontal="center")
            fila_actual += 1

            for e in el:
                periodo = (
                    f"{v(e.fecha_inicio)} → {v(e.fecha_culminacion) or 'Actualidad'}"
                )
                datos_el = [
                    v(e.tipo_institucion),
                    v(e.nombre_entidad),
                    v(e.cargo),
                    periodo,
                ]
                for ci, val in enumerate(datos_el, 1):
                    c = ws2.cell(row=fila_actual, column=ci, value=val)
                    c.font = Font(size=9)
                    c.border = borde_fino()
                fila_actual += 1

        # ── EXPERIENCIA DOCENTE ────────────────────────
        if ed:
            ws2.merge_cells(f"A{fila_actual}:D{fila_actual}")
            sh = ws2.cell(
                row=fila_actual, column=1, value=f"EXPERIENCIA DOCENTE ({len(ed)})"
            )
            estilo_subheader(sh)
            fila_actual += 1

            headers_ed = ["Entidad", "Categoría", "Doc. Acredita", "Período"]
            for ci, h in enumerate(headers_ed, 1):
                c = ws2.cell(row=fila_actual, column=ci, value=h)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="64748B")
                c.border = borde_fino()
                c.alignment = Alignment(horizontal="center")
            fila_actual += 1

            for e in ed:
                periodo = (
                    f"{v(e.fecha_inicio)} → {v(e.fecha_culminacion) or 'Actualidad'}"
                )
                datos_ed = [
                    v(e.nombre_entidad),
                    v(e.categoria),
                    v(e.documento_acredita),
                    periodo,
                ]
                for ci, val in enumerate(datos_ed, 1):
                    c = ws2.cell(row=fila_actual, column=ci, value=val)
                    c.font = Font(size=9)
                    c.border = borde_fino()
                fila_actual += 1

        # ── Espacio entre personas ─────────────────────
        fila_actual += 1

    # Ajustar anchos de columnas Hoja 2
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 25

    # ── Guardar en buffer y devolver ───────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha_hoy = datetime.now().strftime("%Y%m%d")
    nombre_archivo = f"personal_unamba_{fecha_hoy}.xlsx"

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )
